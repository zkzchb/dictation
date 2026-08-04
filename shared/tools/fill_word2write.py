"""fill_word2write.py —— 从 vocab 表回填 word2write 的组词。

规则（与用户约定）：
  * 每个生字(basic_word)从**同一课**的 vocab 词里找包含该字的词填入；
  * 一个 vocab 词只用一次（全局去重）；重复命中时只填第一次出现的生字，
    后面的生字该词留空，等待人工补充——绝不重复填同一个词；
  * 两轮分配：先给所有生字填 word1，再给还能匹配到第二个词的生字填 word2，
    保证"每字先有一个词"再考虑第二个；
  * 拼音直接沿用 vocab 的 vpy（vocab 缺拼音则留空，后续统一补）；
  * 已经手工填好的 word1/word2 保留不动，并把该词计入"已用"。

用法：
  python shared/tools/fill_word2write.py [path/to/wordlist.xlsx]
默认读写 项目根目录的 wordlist_template.xlsx，写前自动备份 .bak。
"""
import sys, os, shutil
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(HERE, "..", "..", "wordlist_template.xlsx")


def col_map(ws):
    """读表头，返回 {列名: 列号(1起)}。"""
    m = {}
    for cell in ws[1]:
        if cell.value is not None:
            m[str(cell.value).strip()] = cell.column
    return m


def is_int(v):
    try:
        int(str(v).strip()); return True
    except (ValueError, TypeError):
        return False


def main():
    path = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
    if not os.path.exists(path):
        print(f"[ERR] 找不到文件: {path}"); sys.exit(1)

    bak = path + ".bak"
    shutil.copy(path, bak)
    print(f"[OK] 已备份 -> {bak}")

    wb = load_workbook(path)
    ws_v = wb["vocab"]
    ws_w = wb["word2write"]

    vm = col_map(ws_v)
    wm = col_map(ws_w)

    # 1. 收集 vocab：按 lid 分组，保持行序 [(vword, vpy)]
    vocab_by_lid = {}
    for row in ws_v.iter_rows(min_row=2):
        vid = row[vm["vid"] - 1].value
        if not is_int(vid):
            continue
        lid = str(row[vm["lid"] - 1].value).strip()
        vword = row[vm["vword"] - 1].value
        vpy = row[vm["vpy"] - 1].value
        if not vword:
            continue
        vocab_by_lid.setdefault(lid, []).append(
            (str(vword).strip(), (str(vpy).strip() if vpy else ""))
        )

    # 2. 收集 word2write 行（行号 + 字 + lid），并记录已手填的词
    w_rows = []
    used = set()  # 全局已用 vword
    for row in ws_w.iter_rows(min_row=2):
        wid = row[wm["wid"] - 1].value
        if not is_int(wid):
            continue
        rn = row[0].row
        char = row[wm["basic_word"] - 1].value
        lid = str(row[wm["lid"] - 1].value).strip()
        w1 = row[wm["word1"] - 1].value
        w2 = row[wm["word2"] - 1].value
        if w1:
            used.add(str(w1).strip())
        if w2:
            used.add(str(w2).strip())
        w_rows.append({
            "rn": rn, "char": (str(char).strip() if char else ""), "lid": lid,
            "has1": bool(w1), "has2": bool(w2),
        })

    def pick(char, lid):
        """返回同课未用、包含 char 的第一个 (vword, vpy)，并标记已用。"""
        for vword, vpy in vocab_by_lid.get(lid, []):
            if vword in used:
                continue
            if char and char in vword:
                used.add(vword)
                return vword, vpy
        return None

    fill1 = fill2 = 0

    # 第一轮：填 word1（跳过已手填的）
    for w in w_rows:
        if w["has1"]:
            continue
        got = pick(w["char"], w["lid"])
        if got:
            ws_w.cell(w["rn"], wm["word1"]).value = got[0]
            if got[1]:
                ws_w.cell(w["rn"], wm["word1py"]).value = got[1]
            w["has1"] = True
            fill1 += 1

    # 第二轮：填 word2（前提是 word1 已有、word2 尚空）
    for w in w_rows:
        if w["has2"] or not w["has1"]:
            continue
        got = pick(w["char"], w["lid"])
        if got:
            ws_w.cell(w["rn"], wm["word2"]).value = got[0]
            if got[1]:
                ws_w.cell(w["rn"], wm["word2py"]).value = got[1]
            fill2 += 1

    wb.save(path)

    total = len(w_rows)
    still_blank = sum(1 for w in w_rows if not w["has1"])
    print(f"[OK] word2write 共 {total} 字")
    print(f"     word1 新填 {fill1}，word2 新填 {fill2}")
    print(f"     仍无任何组词的字: {still_blank}（需人工补充）")
    print(f"     vocab 词已用 {len(used)} 个")


if __name__ == "__main__":
    main()
