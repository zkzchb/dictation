# -*- coding: utf-8 -*-
"""给 word2write / polyphonic / vocab / typo 末尾追加 lid、ltitle、lname 三列，
便于人工核对「该条目所属课程」是否正确。

ID 约定：lid = id // 100
  word2write wid=311101 -> 3111
  polyphonic ppid=311281 -> 3112
  vocab      vid=311151 -> 3111
  typo       tid=311001 -> 3110（单元复习课）

重复运行安全：已存在的三列会被覆盖，不会重复追加。
用法：_build_venv/Scripts/python.exe tools/add_lesson_cols.py
"""
import sys, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "wordlist_template.xlsx")
HELPERS = ["lid", "ltitle", "lname"]
# 各表的 ID 列名
TARGETS = {"word2write": "wid", "polyphonic": "ppid", "vocab": "vid", "typo": "tid"}

wb = openpyxl.load_workbook(XLSX)

# 1. 读 lesson 表建映射
ls = wb["lesson"]
lesson = {}
for r in range(2, ls.max_row + 1):
    lid = ls.cell(r, 1).value
    if isinstance(lid, int):
        lesson[lid] = (ls.cell(r, 2).value, ls.cell(r, 3).value)
print(f"lesson 表: {len(lesson)} 课")

report = []
for sheet, idcol in TARGETS.items():
    ws = wb[sheet]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 定位追加起始列：若已有 helper 列则复用，否则接在末尾
    if hdr[-3:] == HELPERS:
        base = ws.max_column - 3
    else:
        base = ws.max_column
    for i, h in enumerate(HELPERS):
        ws.cell(1, base + i + 1, h)

    bad, filled = [], 0
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None or not isinstance(raw, int):
            continue          # 跳过空行/注释行
        lid = raw // 100
        ws.cell(r, base + 1, lid)
        if lid in lesson:
            t, n = lesson[lid]
            ws.cell(r, base + 2, t)
            ws.cell(r, base + 3, n)
        else:
            ws.cell(r, base + 2, "!! 无此课")
            ws.cell(r, base + 3, None)
            bad.append((r, raw, lid))
        filled += 1
    report.append((sheet, filled, bad))
    print(f"{sheet:12s} 填充 {filled} 行，未匹配 {len(bad)} 行")

wb.save(XLSX)
print("\n已保存。")

# 2. 未匹配明细
print("\n===== 未匹配明细（lid 不在 lesson 表中）=====")
any_bad = False
for sheet, _, bad in report:
    if bad:
        any_bad = True
        print(f"\n[{sheet}]")
        for r, raw, lid in bad:
            print(f"  Excel 第{r}行  id={raw}  推出 lid={lid}")
if not any_bad:
    print("（无。所有条目的 lid 均能在 lesson 表中找到）")
